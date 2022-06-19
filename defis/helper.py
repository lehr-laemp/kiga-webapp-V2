# -*- coding: utf-8 -*-

"""
Begonnen am 12.03.2022

@author: HM

helper.py

    - alle nötigen Funktionen

"""
# ---------------------------------------------------------
import datetime
import os
import pickle
import shutil
import smtplib  # für Mail
import ssl      # für Mail
from email.message import EmailMessage
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
# from threading import Timer
import pyAesCrypt
import streamlit as st
import openpyxl


# ---------------------------------------------------------
def excel_tabelle_entschluesseln():

    print(datetime.datetime.now().strftime('%H-%M-%S'), ': Datenbank entschlüsseln')

    pyAesCrypt.decryptFile('Daten/sus.aes', 'Daten/sus.xlsx', st.secrets['tabelle_passwort'])

    # Mache ein Backup der Datenbank
    # Viele Backups-Dateien
    # backup_dir = 'Backup/' + datetime.datetime.now().strftime('%y-%m-%d-%H-%M-%S') + '-sus.aes'
    # nur 1 Backup-Datei
    backup_dir = 'Backup/backup-sus.aes'
    shutil.copyfile('Daten/sus.aes', backup_dir)

    return True


# ---------------------------------------------------------
def excel_tabelle_in_liste_speichern():

    print(datetime.datetime.now().strftime('%H-%M-%S'), ': Excel-Tabelle in SuS-Liste speichern')

    # öffne und aktiviere Excel-Tabelle
    kiga_datei = openpyxl.load_workbook('Daten/sus.xlsx')
    kiga_tabelle = kiga_datei.active

    sus_liste = []
    sus_einzel = []

    for reihe in range(2, kiga_tabelle.max_row + 1):
        for spalte in range(1, 17):
            sus_einzel.append(
                kiga_tabelle.cell(row=reihe, column=spalte).value)
                
            if spalte == 16:
                sus_liste.append(sus_einzel)
                sus_einzel = []

    # OK print(sus_liste)
    return sus_liste


# ---------------------------------------------------------
def liste_in_pickle_speichern(sus_liste):
    
    print(datetime.datetime.now().strftime('%H-%M-%S'), ': Liste in Pickle speichern')

    # Liste in Pickle speichern
    with open('Daten/sus.tmp', 'wb') as datei_handler:
        pickle.dump(sus_liste, datei_handler)

    return True


# ---------------------------------------------------------
def excel_tabelle_loeschen():
    
    print(datetime.datetime.now().strftime('%H-%M-%S'), ': Excel-Tabelle löschen')

    os.remove('Daten/sus.xlsx')

    return True


# ---------------------------------------------------------
def pickle_in_excel_speichern():
    """
    Speichert Pickle-Dump in Excel-Tabelle
        Öffnet dazu die verschlüsselte Tabelle
        Schreibt die Daten in die Tabelle
        Verschlüsselt die Tabelle wieder
        Löscht die entschlüsselte Tabelle
        Löscht den Pickle-Dump
    """
    
    print(datetime.datetime.now().strftime('%H-%M-%S'), ': Pickle in Excel-Tabelle speichern')

    # Pickle-Dump auslesen
    with open('Daten/sus.tmp', 'rb') as datei_handler:
            sus_liste = pickle.load(datei_handler)

    # Tabelle entschlüsseln
    excel_tabelle_entschluesseln()

    # öffne und aktiviere Excel-Tabelle
    kiga_datei = openpyxl.load_workbook('Daten/sus.xlsx')
    kiga_tabelle = kiga_datei.active

    # Schreibe die SuS-Liste in der Excel-Tabelle
    for reihe in range(len(sus_liste)):
        # OK print(liste[reihe])
        for spalte in range(0, 16):
            kiga_tabelle.cell(row=reihe+2, column=spalte+1).value = sus_liste[reihe][spalte]

    # Speichere die Excel-Tabelle
    kiga_datei.save('Daten/sus.xlsx')

    excel_tabelle_verschluesseln() # und löschen

    # Pickle-Dump löschen
    os.remove('Daten/sus.tmp')

    return True


# ---------------------------------------------------------
def excel_tabelle_verschluesseln():
    
    print(datetime.datetime.now().strftime('%H-%M-%S'), ': Excel-Tabelle verschlüsseln')

    # Excel-Tabelle verschlüsseln
    pyAesCrypt.encryptFile('Daten/sus.xlsx', 'Daten/sus.aes', st.secrets['tabelle_passwort'])

    # Excel-Tabelle löschen
    os.remove('Daten/sus.xlsx')

    # Mache ein Backup der Datenbank
    # backup_dir = 'Backup/' + datetime.datetime.now().strftime('%y-%m-%d-%H-%M-%S') + '-sus.aes'
    backup_dir = 'Backup/backup-sus.aes'
    shutil.copyfile('Daten/sus.aes', backup_dir)
    
    return True


# ---------------------------------------------------------
def liste_aus_pickle_holen():
    
    print(datetime.datetime.now().strftime('%H-%M-%S'),': Liste aus Pickle-Dump holen')

    # Pickle-Dump auslesen
    with open('Daten/sus.tmp', 'rb') as datei_handler:
            sus_liste = pickle.load(datei_handler) 

    return sus_liste
    

# ---------------------------------------------------------
def kiga_standorte_lesen(sus_liste):
    """
    Lies aus der Liste der SuS die möglichen Kiga-Standorte
    return: Liste der Kiga, sortiert
    """

    print(datetime.datetime.now().strftime('%H-%M-%S'), ': Kiga-Standorte einlesen', )

    kiga_liste = []
    for i in range(len(sus_liste)):
        # OK print(sus_liste[i][4])
        kiga_liste.append(sus_liste[i][4])

    return sorted(set(kiga_liste))


# ---------------------------------------------------------
def mail_senden(betreff):
    """
    Schickt eine Nachricht beim Anmelden oder Abmelden
    """

    print(datetime.datetime.now().strftime('%H-%M-%S'), ': Schicke ein Mail')

    # wieviele Backup-Dateien hat es?
    backup_zaehler = 0
    for pfad in os.listdir('Backup/'):
        backup_zaehler += 1
    # OK print('Anzahl Backup-Dateien:', backup_zaehler)

    nachricht = f"""Mail von der Daten-Eingabe:
    
    Eine {betreff}.
    
    Es sind {backup_zaehler} Dateien im Backup-Ordner.
    
    Herzliche Grüsse :-)
    """

    # Angaben für den Server
    gmx_smpt = 'mail.gmx.net'
    gmx_passwort = st.secrets['mail_passwort']
    gmx_port = 587

    # Angaben zum Mail
    mail_von = 'satipati@gmx.ch'
    mail_fuer = 'klameflu@gmail.com'
    mail_betreff = betreff
    mail_text = nachricht

    # Anhang für Mail
    dateiname = 'sus.aes'
    with open('Daten/sus.aes', "rb") as attachment:
        # Add file as application/octet-stream
        # Email client can usually download this automatically as attachment
        part = MIMEBase("application", "octet-stream")
        part.set_payload(attachment.read())

    # Encode file in ASCII characters to send by email    
    encoders.encode_base64(part)

    # Add header as key/value pair to attachment part
    part.add_header(
        "Content-Disposition",
        f"attachment; filename={dateiname}",)
    
    # übersetzen in Email-Format
    nachricht = MIMEMultipart() #EmailMessage()
    # nachricht.set_content(mail_text)

    nachricht['Subject'] = mail_betreff
    nachricht['From'] = mail_von 
    nachricht['To'] = mail_fuer

    # Add attachment to message and convert message to string
    nachricht.attach(part)

    # Verbindung mit Server
    context = ssl.create_default_context()
    try:
        server = smtplib.SMTP(gmx_smpt, gmx_port)
        #server.set_debuglevel(1)
        server.starttls(context=context)
        server.login(mail_von, gmx_passwort)
        server.send_message(nachricht)
    except Exception as e:
        # print(e)
        st.warning('Kann Email nicht senden.')
    finally:
        server.quit()


# ---------------------------------------------------------
def melde_status():

    print(datetime.datetime.now().strftime('%H-%M-%S'),': Melde Status')

    global timer
    
    # ist jemand angemeldet?
    anmelde_status = 2 #st.session_state['angemeldet']
    print(anmelde_status)
    # geht nicht print(st.session_state.angemeldet)
    # geht nicht print(st.session_state['timer'])

    # wieviele Backup-Dateien hat es?
    zaehler = 0
    for pfad in os.listdir('Backup/'):
        zaehler += 1
    print('Anzahl Backup-Dateien:', zaehler)

    # Timer stoppen und neu starten
    timer.cancel()
    start_timer()


# ---------------------------------------------------------
def start_timer():

    global timer

    print(datetime.datetime.now().strftime('%H-%M-%S'),': Starte den Timer')

    # ti = Timer(30, melde_automatisch_ab, args=None)
    timer = Timer(10, melde_status, args=None)
    timer.start()


# ---------------------------------------------------------
def stop_timer():

    global timer

    print(datetime.datetime.now().strftime('%H-%M-%S'),': Stoppe den Timer')

    timer.cancel()

